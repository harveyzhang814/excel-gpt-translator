from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QProgressBar, QListWidget,
    QFileDialog, QMessageBox, QComboBox, QCheckBox,
    QLineEdit, QDialog, QFormLayout, QTextEdit,
    QListWidgetItem, QListWidget, QFrame
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from core.config import Config
from core.translator import Translator
from .dialogs.task_dialog import TaskDialog
from .dialogs.settings_dialog import SettingsDialog
from .widgets.task_widget import TaskWidget
import pandas as pd
import uuid
from openpyxl import load_workbook
import re

class TranslationThread(QThread):
    progress_updated = pyqtSignal(int)
    finished = pyqtSignal()
    error = pyqtSignal(str)

    def __init__(self, translator, task_data):
        super().__init__()
        self.translator = translator
        self.task_data = task_data
        # Connect translator's progress signal to our progress signal
        self.translator.progress_updated.connect(self.progress_updated.emit)

    def run(self):
        try:
            print("\n=== Starting Translation Task ===")
            print(f"Task data: {self.task_data}")
            self.translator.translate_excel(self.task_data)
            print("=== Translation Task Completed ===\n")
            self.finished.emit()
        except Exception as e:
            print(f"\n=== Translation Task Failed ===")
            print(f"Error details: {str(e)}")
            print(f"Task data: {self.task_data}")
            print("===============================\n")
            self.error.emit(str(e))

class MainWindow(QMainWindow):
    def __init__(self, config: Config):
        super().__init__()
        self.config = config
        self.translator = Translator(config)
        self.tasks = {}  # Dictionary of task_id -> task_data
        self.translation_threads = {}  # Dictionary of task_id -> thread
        
        self.setWindowTitle("Excel GPT Translator")
        self.setMinimumSize(800, 600)
        
        # Create central widget and layout
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)
        
        # Create toolbar
        toolbar_layout = QHBoxLayout()
        self.create_task_btn = QPushButton("Create Task")
        self.settings_btn = QPushButton("Settings")
        toolbar_layout.addWidget(self.create_task_btn)
        toolbar_layout.addWidget(self.settings_btn)
        toolbar_layout.addStretch()
        layout.addLayout(toolbar_layout)
        
        # Create task list
        self.task_list = QListWidget()
        layout.addWidget(self.task_list)
        
        # Connect signals
        self.create_task_btn.clicked.connect(self.create_task)
        self.settings_btn.clicked.connect(self.show_settings)
    
    def create_task(self):
        dialog = TaskDialog(self.config, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            task_data = dialog.get_task_data()
            task_id = str(uuid.uuid4())
            self.tasks[task_id] = task_data
            self.update_task_list()
    
    def edit_task(self, task_id):
        # Get current task data
        task_data = self.tasks[task_id]
        
        # Create dialog with current values
        dialog = TaskDialog(self.config, self)
        dialog.file_path.setText(task_data['file'])
        dialog._update_sheet_selector(task_data['file'])
        dialog.sheet_selector.setCurrentText(task_data['sheet'])
        if task_data.get('cell_range'):
            dialog.cell_range.setText(task_data['cell_range'])
        dialog.current_lang.setCurrentText(task_data['current_language'])
        dialog.target_langs.setCurrentText(task_data['target_languages'][0])
        dialog.comparison_mode.setChecked(task_data['comparison_mode'])
        dialog.prompt_text.setText(task_data['prompt'])
        if task_data.get('field'):
            dialog.field_input.setText(task_data['field'])
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Update task data
            self.tasks[task_id] = dialog.get_task_data()
            self.update_task_list()
    
    def show_settings(self):
        dialog = SettingsDialog(self.config, self)
        dialog.exec()
    
    def update_task_list(self):
        self.task_list.clear()
        for task_id, task_data in self.tasks.items():
            item = QListWidgetItem()
            widget = TaskWidget(task_id, task_data)
            
            # Connect signals
            widget.start_btn.clicked.connect(lambda checked, t_id=task_id: self.start_translation(t_id))
            widget.edit_btn.clicked.connect(lambda checked, t_id=task_id: self.edit_task(t_id))
            widget.remove_btn.clicked.connect(lambda checked, t_id=task_id: self.remove_task(t_id))
            
            self.task_list.addItem(item)
            self.task_list.setItemWidget(item, widget)
            item.setSizeHint(widget.sizeHint())
    
    def start_translation(self, task_id):
        # Find the task widget
        for i in range(self.task_list.count()):
            item = self.task_list.item(i)
            widget = self.task_list.itemWidget(item)
            if widget.task_id == task_id:
                # Disable start button
                widget.start_btn.setEnabled(False)
                widget.edit_btn.setEnabled(False)
                
                # Create and start translation thread
                thread = TranslationThread(self.translator, self.tasks[task_id])
                thread.progress_updated.connect(widget.progress_bar.setValue)
                thread.finished.connect(lambda w=widget: self.on_translation_finished(w))
                thread.error.connect(lambda msg, w=widget: self.on_translation_error(msg, w))
                
                self.translation_threads[task_id] = thread
                thread.start()
                break
    
    def remove_task(self, task_id):
        if task_id in self.translation_threads:
            self.translation_threads[task_id].terminate()
            del self.translation_threads[task_id]
        
        del self.tasks[task_id]
        self.update_task_list()
    
    def on_translation_finished(self, widget):
        widget.start_btn.setEnabled(True)
        widget.edit_btn.setEnabled(True)
        widget.progress_bar.setValue(100)
        QMessageBox.information(self, "Success", "Translation completed successfully!")
    
    def on_translation_error(self, error_msg, widget):
        widget.start_btn.setEnabled(True)
        widget.edit_btn.setEnabled(True)
        QMessageBox.critical(self, "Error", f"Translation failed: {error_msg}")

class TaskDialog(QDialog):
    def __init__(self, config: Config, parent=None):
        super().__init__(parent)
        self.config = config
        self.setWindowTitle("Create Translation Task")
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # File selection
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("Excel File:"))
        self.file_path = QLineEdit()
        self.file_path.setReadOnly(True)
        file_layout.addWidget(self.file_path)
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)
        
        # Sheet selection
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(QLabel("Sheet:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setEnabled(False)
        sheet_layout.addWidget(self.sheet_selector)
        layout.addLayout(sheet_layout)
        
        # Cell range
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("Cell Range (required, e.g., A1:B4):"))
        self.cell_range = QLineEdit()
        self.cell_range.setPlaceholderText("Enter cell range (e.g., A1:B4)")
        self.cell_range.textChanged.connect(self.validate_input)
        range_layout.addWidget(self.cell_range)
        layout.addLayout(range_layout)
        
        # Language selection
        lang_layout = QHBoxLayout()
        lang_layout.addWidget(QLabel("Current Language:"))
        self.current_lang = QComboBox()
        self.current_lang.addItems(self.config.get_supported_languages())
        lang_layout.addWidget(self.current_lang)
        
        lang_layout.addWidget(QLabel("Target Language:"))
        self.target_langs = QComboBox()
        self.target_langs.addItems(self.config.get_supported_languages())
        lang_layout.addWidget(self.target_langs)
        
        # Field/Industry (Optional)
        field_layout = QHBoxLayout()
        field_layout.addWidget(QLabel("Field/Industry (Optional):"))
        self.field_input = QLineEdit()
        self.field_input.setPlaceholderText("e.g., Medical, Legal, Technical, etc.")
        field_layout.addWidget(self.field_input)
        layout.addLayout(field_layout)
        
        # Comparison mode
        self.comparison_mode = QCheckBox("Enable comparison mode")
        layout.addWidget(self.comparison_mode)
        
        # Prompt
        prompt_layout = QVBoxLayout()
        prompt_header = QHBoxLayout()
        prompt_header.addWidget(QLabel("Translation Prompt:"))
        reset_prompt_btn = QPushButton("Reset to Default")
        reset_prompt_btn.clicked.connect(self.reset_prompt)
        prompt_header.addWidget(reset_prompt_btn)
        prompt_layout.addLayout(prompt_header)
        
        # Help text for prompt placeholders
        help_label = QLabel(
            "Required placeholders:\n"
            "{current_lang} - Source language\n"
            "{target_lang} - Target language\n"
            "{text} - Text to translate"
        )
        help_label.setStyleSheet("color: gray; font-size: 10pt;")
        prompt_layout.addWidget(help_label)
        
        self.prompt_text = QTextEdit()
        self.prompt_text.setPlainText(self.config.get_default_prompt())
        self.prompt_text.textChanged.connect(self.validate_prompt)
        prompt_layout.addWidget(self.prompt_text)
        
        # Warning label for invalid prompt
        self.prompt_warning = QLabel()
        self.prompt_warning.setStyleSheet("color: red;")
        self.prompt_warning.hide()
        prompt_layout.addWidget(self.prompt_warning)
        
        layout.addLayout(prompt_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        self.ok_btn = QPushButton("OK")
        self.ok_btn.clicked.connect(self.validate_and_accept)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)
        
        # Initial validation
        self.validate_input()
    
    def validate_input(self):
        """Validate all required inputs."""
        is_valid = True
        
        # Validate cell range
        cell_range = self.cell_range.text().strip()
        if not cell_range:
            is_valid = False
        elif not self._is_valid_cell_range(cell_range):
            is_valid = False
        
        # Validate file selection
        if not self.file_path.text().strip():
            is_valid = False
        
        # Validate sheet selection
        if not self.sheet_selector.currentText():
            is_valid = False
        
        self.ok_btn.setEnabled(is_valid)
    
    def _is_valid_cell_range(self, cell_range: str) -> bool:
        """Check if the cell range format is valid."""
        pattern = r'^[A-Z]+[1-9][0-9]*:[A-Z]+[1-9][0-9]*$'
        return bool(re.match(pattern, cell_range))
    
    def validate_and_accept(self):
        """Validate all inputs before accepting."""
        if not self.cell_range.text().strip():
            QMessageBox.warning(self, "Validation Error", "Cell range is required.")
            return
        
        if not self._is_valid_cell_range(self.cell_range.text().strip()):
            QMessageBox.warning(self, "Validation Error", "Invalid cell range format. Please use format like 'A1:B4'.")
            return
        
        self.accept()
    
    def reset_prompt(self):
        """Reset the prompt to default value."""
        self.prompt_text.setPlainText(self.config.get_default_prompt())
    
    def validate_prompt(self):
        """Validate that the prompt contains all required placeholders."""
        prompt = self.prompt_text.toPlainText()
        missing = []
        
        if "{current_lang}" not in prompt:
            missing.append("{current_lang}")
        if "{target_lang}" not in prompt:
            missing.append("{target_lang}")
        if "{text}" not in prompt:
            missing.append("{text}")
        
        if missing:
            self.prompt_warning.setText(f"Missing required placeholders: {', '.join(missing)}")
            self.prompt_warning.show()
            self.ok_btn.setEnabled(False)
        else:
            self.prompt_warning.hide()
            self.validate_input()  # Check other inputs as well
    
    def browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            self.file_path.setText(file_name)
            self._update_sheet_selector(file_name)
            self.validate_input()
    
    def _update_sheet_selector(self, file_path):
        try:
            wb = load_workbook(file_path, read_only=True)
            self.sheet_selector.clear()
            self.sheet_selector.addItems(wb.sheetnames)
            self.sheet_selector.setEnabled(True)
            self.validate_input()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to read Excel file: {str(e)}")
            self.sheet_selector.setEnabled(False)
            self.validate_input()
    
    def get_task_data(self):
        return {
            'file': self.file_path.text(),
            'sheet': self.sheet_selector.currentText(),
            'cell_range': self.cell_range.text().strip(),
            'current_language': self.current_lang.currentText(),
            'target_languages': [self.target_langs.currentText()],
            'comparison_mode': self.comparison_mode.isChecked(),
            'prompt': self.prompt_text.toPlainText(),
            'field': self.field_input.text().strip()
        }

class SettingsDialog(QDialog):
    def __init__(self, config: Config, parent=None):
        super().__init__(parent)
        self.config = config
        self.setWindowTitle("Settings")
        self.setup_ui()
    
    def setup_ui(self):
        layout = QFormLayout(self)
        
        # API Key
        self.api_key = QLineEdit()
        self.api_key.setText(self.config.get_api_key())
        layout.addRow("OpenAI API Key:", self.api_key)
        
        # Buttons
        buttons = QHBoxLayout()
        save_btn = QPushButton("Save")
        cancel_btn = QPushButton("Cancel")
        buttons.addWidget(save_btn)
        buttons.addWidget(cancel_btn)
        layout.addRow("", buttons)
        
        save_btn.clicked.connect(self.save_settings)
        cancel_btn.clicked.connect(self.reject)
    
    def save_settings(self):
        self.config.save_api_key(self.api_key.text())
        self.accept() 
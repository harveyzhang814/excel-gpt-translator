from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QComboBox, QCheckBox, QTextEdit, QFileDialog
)
import re
from openpyxl import load_workbook

class TaskDialog(QDialog):
    def __init__(self, config, parent=None):
        super().__init__(parent)
        self.config = config
        self.setWindowTitle("Create Translation Task")
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # File selection
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("Excel File:"))
        self.file_path = QLineEdit()
        self.file_path.setReadOnly(True)
        file_layout.addWidget(self.file_path)
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_file)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)
        
        # Sheet selection
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(QLabel("Sheet:"))
        self.sheet_selector = QComboBox()
        self.sheet_selector.setEnabled(False)
        sheet_layout.addWidget(self.sheet_selector)
        layout.addLayout(sheet_layout)
        
        # Cell range
        range_layout = QHBoxLayout()
        range_layout.addWidget(QLabel("Cell Range (required, e.g., A1:B4):"))
        self.cell_range = QLineEdit()
        self.cell_range.setPlaceholderText("Enter cell range (e.g., A1:B4)")
        self.cell_range.textChanged.connect(self.validate_input)
        range_layout.addWidget(self.cell_range)
        layout.addLayout(range_layout)
        
        # Language selection
        lang_layout = QHBoxLayout()
        lang_layout.addWidget(QLabel("Current Language:"))
        self.current_lang = QComboBox()
        self.current_lang.addItems(self.config.get_supported_languages())
        lang_layout.addWidget(self.current_lang)
        
        lang_layout.addWidget(QLabel("Target Language:"))
        self.target_langs = QComboBox()
        self.target_langs.addItems(self.config.get_supported_languages())
        lang_layout.addWidget(self.target_langs)
        
        # Field/Industry (Optional)
        field_layout = QHBoxLayout()
        field_layout.addWidget(QLabel("Field/Industry (Optional):"))
        self.field_input = QLineEdit()
        self.field_input.setPlaceholderText("e.g., Medical, Legal, Technical, etc.")
        field_layout.addWidget(self.field_input)
        layout.addLayout(field_layout)
        
        # Comparison mode
        self.comparison_mode = QCheckBox("Enable comparison mode")
        layout.addWidget(self.comparison_mode)
        
        # Prompt
        prompt_layout = QVBoxLayout()
        prompt_header = QHBoxLayout()
        prompt_header.addWidget(QLabel("Translation Prompt:"))
        reset_prompt_btn = QPushButton("Reset to Default")
        reset_prompt_btn.clicked.connect(self.reset_prompt)
        prompt_header.addWidget(reset_prompt_btn)
        prompt_layout.addLayout(prompt_header)
        
        # Help text for prompt placeholders
        help_label = QLabel(
            "Required placeholders:\n"
            "{current_lang} - Source language\n"
            "{target_lang} - Target language\n"
            "{text} - Text to translate"
        )
        help_label.setStyleSheet("color: gray; font-size: 10pt;")
        prompt_layout.addWidget(help_label)
        
        self.prompt_text = QTextEdit()
        self.prompt_text.setPlainText(self.config.get_default_prompt())
        self.prompt_text.textChanged.connect(self.validate_prompt)
        prompt_layout.addWidget(self.prompt_text)
        
        # Warning label for invalid prompt
        self.prompt_warning = QLabel()
        self.prompt_warning.setStyleSheet("color: red;")
        self.prompt_warning.hide()
        prompt_layout.addWidget(self.prompt_warning)
        
        layout.addLayout(prompt_layout)
        
        # Buttons
        button_layout = QHBoxLayout()
        self.ok_btn = QPushButton("OK")
        self.ok_btn.clicked.connect(self.validate_and_accept)
        self.cancel_btn = QPushButton("Cancel")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        layout.addLayout(button_layout)
        
        # Initial validation
        self.validate_input()
    
    def validate_input(self):
        """Validate all required inputs."""
        is_valid = True
        
        # Validate cell range
        cell_range = self.cell_range.text().strip()
        if not cell_range:
            is_valid = False
        elif not self._is_valid_cell_range(cell_range):
            is_valid = False
        
        # Validate file selection
        if not self.file_path.text().strip():
            is_valid = False
        
        # Validate sheet selection
        if not self.sheet_selector.currentText():
            is_valid = False
        
        self.ok_btn.setEnabled(is_valid)
    
    def _is_valid_cell_range(self, cell_range: str) -> bool:
        """Check if the cell range format is valid."""
        pattern = r'^[A-Z]+[1-9][0-9]*:[A-Z]+[1-9][0-9]*$'
        return bool(re.match(pattern, cell_range))
    
    def validate_and_accept(self):
        """Validate all inputs before accepting."""
        if not self.cell_range.text().strip():
            QMessageBox.warning(self, "Validation Error", "Cell range is required.")
            return
        
        if not self._is_valid_cell_range(self.cell_range.text().strip()):
            QMessageBox.warning(self, "Validation Error", "Invalid cell range format. Please use format like 'A1:B4'.")
            return
        
        self.accept()
    
    def reset_prompt(self):
        """Reset the prompt to default value."""
        self.prompt_text.setPlainText(self.config.get_default_prompt())
    
    def validate_prompt(self):
        """Validate that the prompt contains all required placeholders."""
        prompt = self.prompt_text.toPlainText()
        missing = []
        
        if "{current_lang}" not in prompt:
            missing.append("{current_lang}")
        if "{target_lang}" not in prompt:
            missing.append("{target_lang}")
        if "{text}" not in prompt:
            missing.append("{text}")
        
        if missing:
            self.prompt_warning.setText(f"Missing required placeholders: {', '.join(missing)}")
            self.prompt_warning.show()
            self.ok_btn.setEnabled(False)
        else:
            self.prompt_warning.hide()
            self.validate_input()  # Check other inputs as well
    
    def browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            self.file_path.setText(file_name)
            self._update_sheet_selector(file_name)
            self.validate_input()
    
    def _update_sheet_selector(self, file_path):
        try:
            wb = load_workbook(file_path, read_only=True)
            self.sheet_selector.clear()
            self.sheet_selector.addItems(wb.sheetnames)
            self.sheet_selector.setEnabled(True)
            self.validate_input()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to read Excel file: {str(e)}")
            self.sheet_selector.setEnabled(False)
            self.validate_input()
    
    def get_task_data(self):
        return {
            'file': self.file_path.text(),
            'sheet': self.sheet_selector.currentText(),
            'cell_range': self.cell_range.text().strip(),
            'current_language': self.current_lang.currentText(),
            'target_languages': [self.target_langs.currentText()],
            'comparison_mode': self.comparison_mode.isChecked(),
            'prompt': self.prompt_text.toPlainText(),
            'field': self.field_input.text().strip()
        } 
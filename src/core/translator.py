import pandas as pd
from openai import OpenAI
from pathlib import Path
import json
from tqdm import tqdm
import re
from PyQt6.QtCore import QObject, pyqtSignal
import os
from openpyxl import load_workbook

class Translator(QObject):
    progress_updated = pyqtSignal(int)
    
    def __init__(self, config):
        super().__init__()
        self.config = config
        self.client = OpenAI(api_key=config.get_api_key())
    
    def _parse_cell_range(self, cell_range: str) -> tuple:
        """Parse the cell range (e.g., "A1:B4") into start and end cell references."""
        try:
            start_cell, end_cell = cell_range.split(':')
            return start_cell, end_cell
        except Exception as e:
            raise ValueError(f"Invalid cell range format. Please use format like 'A1:B4'. Error: {str(e)}")
    
    def _get_column_letter(self, col_idx: int) -> str:
        """Convert 0-based column index to Excel column letter (A, B, C, ..., Z, AA, AB, etc.)."""
        result = ""
        while col_idx >= 0:
            col_idx, remainder = divmod(col_idx, 26)
            result = chr(65 + remainder) + result
            col_idx -= 1
        return result
    
    def _get_cell_reference(self, row_idx: int, col_idx: int) -> str:
        """Convert 0-based row and column indices to Excel cell reference (e.g., 'A1')."""
        col_letter = self._get_column_letter(col_idx)
        return f"{col_letter}{row_idx + 1}"

    def translate_excel(self, task_data):
        """Translate an Excel file according to the task settings."""
        self.task_data = task_data  # Store task data for use in _translate_text
        file_path = task_data['file']
        sheet_name = task_data['sheet']
        cell_range = task_data.get('cell_range')
        if not cell_range:
            raise ValueError("Cell range is required. Please specify a range (e.g., 'A1:B4')")
            
        current_lang = task_data['current_language']
        target_langs = task_data['target_languages']
        comparison_mode = task_data['comparison_mode']
        prompt_template = task_data.get('prompt', 
            "Please translate the following text from {current_lang} to {target_lang}:\n\n{text}"
        )
        
        try:
            # Read Excel file using openpyxl to preserve formatting
            wb = load_workbook(file_path)
            sheet = wb[sheet_name]
            
            # Parse the cell range to get start and end cells
            start_cell, end_cell = self._parse_cell_range(cell_range)
            
            # Get the cells to translate - only those with text content
            cells_to_translate = []
            for row in sheet[start_cell:end_cell]:
                for cell in row:
                    if self._should_translate_cell(cell):
                        cells_to_translate.append(cell)
            
            print(f"Found {len(cells_to_translate)} cells with text content to translate")
            
            # Calculate total cells for progress tracking
            total_cells = len(cells_to_translate) * len(target_langs)
            processed_cells = 0
            
            # Process each target language
            for target_lang in target_langs:
                # Create a new workbook for this translation
                output_path = self._get_output_path(file_path, target_lang)
                
                # Copy the original workbook
                wb.save(output_path)
                new_wb = load_workbook(output_path)
                new_sheet = new_wb[sheet_name]
                
                # Translate each cell
                for cell in cells_to_translate:
                    try:
                        # Get cell text
                        cell_str = self._get_cell_text(cell)
                        print(f"\nProcessing cell {cell.coordinate} with value: {cell_str}")
                        
                        # Translate the cell content
                        translated_text = self._translate_text(
                            cell_str, current_lang, target_lang, prompt_template
                        )
                        
                        # Get the target cell in the new workbook
                        target_cell = new_sheet[cell.coordinate]
                        
                        if comparison_mode:
                            # Format with original and translated text
                            target_cell.value = f"{cell_str}\n\n{translated_text}"
                            target_cell.alignment = target_cell.alignment.copy(wrap_text=True)
                        else:
                            target_cell.value = translated_text
                        
                        # Update progress
                        processed_cells += 1
                        progress = int((processed_cells / total_cells) * 100)
                        self.progress_updated.emit(progress)
                        
                    except Exception as e:
                        print(f"Error processing cell {cell.coordinate}: {str(e)}")
                        raise
                
                # Save the translated workbook
                new_wb.save(output_path)
                
        except Exception as e:
            print(f"Error in translate_excel: {str(e)}")
            raise
    
    def _should_translate_cell(self, cell):
        """Determine if a cell should be translated based on its content."""
        # Skip empty cells
        if cell.value is None:
            return False
        
        # Skip cells with numbers only
        if isinstance(cell.value, (int, float)):
            return False
        
        # Skip cells with formulas (formula results are shown as values)
        if cell.data_type == 'f':
            return False
        
        # Skip cells with only whitespace, numbers, or special characters
        if isinstance(cell.value, str):
            # Strip whitespace
            text = cell.value.strip()
            if not text:
                return False
            
            # Skip if only numbers and punctuation
            if re.match(r'^[\d\s\.\,\:\;\-\+\=\(\)\[\]\{\}\/\\\|\!@#\$%\^&\*]*$', text):
                return False
            
            # Skip if too short (like "a", "123", etc.)
            if len(text) < 2:
                return False
            
            return True
        
        return False
    
    def _get_cell_text(self, cell):
        """Extract text content from a cell."""
        if isinstance(cell.value, str):
            return cell.value.strip()
        return str(cell.value)
    
    def _translate_dataframe(self, df, current_lang, target_lang, comparison_mode, prompt_template, processed_cells, total_cells):
        """Translate a pandas DataFrame."""
        translated_df = pd.DataFrame(index=df.index, columns=df.columns)
        
        try:
            # Process each cell
            for col_idx, col in enumerate(df.columns):
                for row_idx in range(len(df)):
                    try:
                        cell_value = df.iloc[row_idx, col_idx]
                        # Convert to string and check if it's not empty
                        if pd.notna(cell_value):
                            # Convert numbers to strings with proper formatting
                            if isinstance(cell_value, (int, float)):
                                if isinstance(cell_value, int) or cell_value.is_integer():
                                    cell_str = str(int(cell_value))
                                else:
                                    cell_str = str(float(cell_value))
                            else:
                                cell_str = str(cell_value).strip()

                            if cell_str:
                                try:
                                    print(f"\nProcessing cell [{row_idx}, {col}] with value: {cell_str} (type: {type(cell_value)})")
                                    translated_text = self._translate_text(
                                        cell_str, current_lang, target_lang, prompt_template
                                    )
                                    translated_df.iloc[row_idx, col_idx] = translated_text
                                except Exception as e:
                                    print(f"Translation error for cell [{row_idx}, {col}] with value '{cell_str}': {str(e)}")
                                    raise Exception(f"Failed to translate cell [{row_idx}, {col}] with value '{cell_str}': {str(e)}")
                    except Exception as e:
                        print(f"Error processing cell [{row_idx}, {col}]: {str(e)}")
                        raise
                    
                    # Update progress after each cell
                    current_cell = (col_idx * len(df)) + row_idx + 1
                    current_progress = ((processed_cells + current_cell) / total_cells) * 100
                    self.progress_updated.emit(int(current_progress))
            
            return translated_df
        except Exception as e:
            print(f"Error in _translate_dataframe: {str(e)}")
            raise
    
    def _translate_text(self, text: str, current_lang: str, target_lang: str, prompt_template: str) -> str:
        """Translate text using GPT API."""
        try:
            # Create a more specific system message
            system_message = "You are a professional translator. Your task is to translate text while preserving meaning and tone. Only respond with the translated text, no explanations or additional content."
            
            # Format the prompt using the Config class's format_prompt method if using default prompt
            if prompt_template == self.config.get_default_prompt():
                user_prompt = self.config.format_prompt(
                    current_lang=current_lang,
                    target_lang=target_lang,
                    text=text,
                    field=self.task_data.get('field', '')
                )
            else:
                # Use custom prompt template
                try:
                    user_prompt = prompt_template.format(
                        current_lang=current_lang,
                        target_lang=target_lang,
                        text=text
                    )
                except KeyError as e:
                    print(f"Warning: Prompt template missing placeholders. Using fallback template.")
                    fallback_template = "Please translate the following text from {current_lang} to {target_lang}:\n\n{text}"
                    user_prompt = fallback_template.format(
                        current_lang=current_lang,
                        target_lang=target_lang,
                        text=text
                    )
            
            print("\n=== Translation Request ===")
            print(f"Source text: '{text}'")
            print(f"From language: {current_lang}")
            print(f"To language: {target_lang}")
            print(f"System message: {system_message}")
            print(f"User prompt template: {prompt_template}")
            print(f"Formatted user prompt: {user_prompt}")
            print("========================\n")
            
            response = self.client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {
                        "role": "system",
                        "content": system_message
                    },
                    {
                        "role": "user",
                        "content": user_prompt
                    }
                ]
            )
            result = response.choices[0].message.content.strip()
            
            print("\n=== Translation Result ===")
            print(f"Original: '{text}'")
            print(f"Translated: '{result}'")
            print("========================\n")
            
            return result
        except Exception as e:
            print("\n=== Translation Error ===")
            print(f"Error type: {type(e).__name__}")
            print(f"Error message: {str(e)}")
            print(f"Input text: '{text}'")
            print(f"Current language: {current_lang}")
            print(f"Target language: {target_lang}")
            print(f"Prompt template: {prompt_template}")
            print("========================\n")
            raise Exception(f"Translation failed: {str(e)}")
    
    def _get_output_path(self, input_path, target_lang):
        """Generate output file path."""
        path = Path(input_path)
        return path.parent / f"{path.stem}_{target_lang}{path.suffix}" 